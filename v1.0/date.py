import os
import re
import openpyxl as px

# 获取当前脚本所在目录的路径
script_directory = os.path.dirname(os.path.abspath(__file__))

# 定义输出文件夹路径
output_directory = os.path.join(script_directory, 'output')

# 创建输出文件夹（如果不存在）
os.makedirs(output_directory, exist_ok=True)

# 遍历脚本目录及其子目录中的所有.log文件
log_files = []
for root, dirs, files in os.walk(script_directory):
    for file in files:
        if file.endswith('.log'):
            log_files.append(os.path.join(root, file))

# 创建一个新的Excel工作簿
wb = px.Workbook()
ws = wb.active

# 在Excel表格中添加表头
ws['A1'] = 'Distance'
ws['B1'] = 'Power(mW)'

# 循环处理每个日志文件
for log_file_path in log_files:
    # 从文件路径中提取不包括扩展名的部分作为文件名
    base_name = os.path.splitext(os.path.basename(log_file_path))[0]

    # 打开日志文件
    with open(log_file_path, 'r') as log_file:
        # 读取文件内容
        lines = log_file.readlines()

        # 初始化距离和Power(mW)
        distance = None
        power_mW = None

        # 遍历文件的每一行
        for line in lines:
            # 使用正则表达式匹配"Aver"后的第一个小数
            match = re.search(r'Aver\s+(\d+\.\d+)', line)
            if match:
                power_mW = float(match.group(1))
                break  # 找到后即可退出循环

        # 将数据写入Excel表格
        row = [base_name, power_mW]
        ws.append(row)

# 保存Excel文件
excel_file_path = os.path.join(script_directory, 'output/log_data.xlsx')
wb.save(excel_file_path)